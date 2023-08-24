import glob as g
from signal import SIG_DFL
from openpyxl import Workbook
from openpyxl import load_workbook
import re
import os
import nltk
nltk.download('wordnet')
from utilities.buying_signal_utilities import buying_signal_utilities
from tqdm import tqdm
bs_utils = buying_signal_utilities()

wb = Workbook()
ws1 = wb.active
wb_filename = "buying_signal_linkedin.xlsx"

ws1.append(['Keyword', 'Url', 'Post Content', 'Comment', 'Authors_name', 'Buying_signal', 'Source']) #Header
ws_row = 2 #Row and col indexing start at 1 instead of 0. Row 1 is header row.
ws_col = 1
current_wb = load_workbook(filename=r"C:\Users\HP\Desktop\Orcawise\social_media_data-linkedln-Orcawise_project\Orcawise_project\raw_files\raw_file_linkedin.xlsx")
wb_active = current_wb.active
for row in tqdm(wb_active.iter_rows(min_row=2, min_col=1, max_row=wb_active.max_row, max_col=wb_active.max_column,values_only=True)):
    row = list(row)
    row[0] = row[0].lower()
    row[2] = bs_utils.cleaning_text(row[2])
    row.append(bs_utils.matches(row[2]))
    row.append('linkedin')
    ws1.append(row)

wb.save(wb_filename)
